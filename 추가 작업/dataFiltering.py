from openpyxl import workbook, load_workbook, utils
from openpyxl.styles import Font, Color, Fill
import os
import sys

print("\n=======================================")
print("            데이터 필터링\n")
print("      1. 키워드 포함 데이터 삭제")
print("   2. 키워드 포함 데이터 삭제 + 복사")
print("    3. 키워드 포함 빨간색 하이라이팅\n")
print("=======================================")

option = int(input(">> 옵션번호 입력 (숫자만 입력할 것) : "))

excel_name = "sample.xlsx"

# files = os.listdir()
# for file in files:
#     if os.path.isdir(file):
#         dirFiles = os.listdir(file)
#         if dirFiles.count("Filtering.py") >=1 :
#             for dirFile in dirFiles:
#                 if dirFile.count(".xlsx") >=1 :
#                     print("\n",dirFile,"파일이 자동으로 선택되었습니다.")
#                     choose = input(">> 이 파일로 진행하시겠습니까? (y/n) : ")
#                     if choose == 'Y' or choose == 'y' :
#                         excel_name = dirFile
#                         break

# if excel_name =="" :
#     print("\n!!! 엑셀 파일이 존재하지 않습니다. !!!")
#     exit()

# else:
#     print("\n",excel_name, "에서 데이터 필터링이 진행됩니다.\n")

# prefix = os.path.dirname( os.path.abspath( __file__))
# excel_name = prefix + "/" + excel_name
wb = load_workbook(excel_name, data_only=True)

excel_sheet_name = "Sheet1"
keyword_name = "keyword"
highlight_name = "highlight"

ws = wb[excel_sheet_name]

if option == 2 :
    new_ws_name = "Sheet2"
    if wb.sheetnames.count(new_ws_name)==1 :
        print("생성하려는 'Sheet2' 가 이미 존재합니다.")
        print("다음을 제외한 새로운 Sheet이름을 입력해주세요")
        print("----------------------------------------------")
        print(wb.sheetnames)
        new_ws_name = input(" >> 생성할 Sheet 이름 입력 :  ")
    new_ws = wb.create_sheet()
    new_ws.title = new_ws_name

    for name in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
        new_ws[name+'1'] = ws[name+'1'].value

keyword_name = prefix + "/" + keyword_name + ".txt"
highlight_name = prefix + "/" + highlight_name + ".txt"

#option 1,2 인 경우 - 공통적으로 삭제 진행 후 option 2인 경우 추가 복사 진행
if option == 1 or option == 2 :
    deleteRows = []
    out_cnt = 0
    try :
        keywords = []
        with open(keyword_name,"r") as f:
            for line in f.readlines():
                keywords.append(line.strip())
    except Exception as e:
        print("keyword 텍스트 파일 읽어오기 실패 (이유 : {e}")

    for i, row in enumerate(ws.iter_rows(min_row = 2)):
        ex_flag = False
        if row[0].value is None:
            break
        
        for word in keywords:
            if row[4].value.find(word) != -1:
                ex_flag = True
                break
        
        # 키워드 발견 시 제외
        if ex_flag:
            if option == 2 :
                out_cnt += 1
                #------------- 기본 : A~R열 일 때 ---------------
                new_ws['A' + str(out_cnt + 1)] = out_cnt
                rowNumber = 1
                for name in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
                    new_ws[name+str(out_cnt +1)] = row[rowNumber].value
                    rowNumber += 1

                #---- 최근 구매건수, 리뷰평점 추가될 시 (A~T열 일 때) ----
                #new_ws['S' + str(out_cnt + 1)] = row[18].value
                #new_ws['T' + str(out_cnt + 1)] = row[19].value
            deleteRows.append(i+2)

    deleteRows.reverse()
    for delete in deleteRows :
        ws.delete_rows(delete)

    print("=================================================")
    print("                 필터링 결과")
    print("     총 %d 개의 키워드가 발견되어 삭제됐습니다" %(len(deleteRows)))
    print("=================================================")


# option 3 인 경우
if option == 3 : 
    try :
        keywords = []
        with open(highlight_name,"r") as f:
            for line in f.readlines():
                keywords.append(line.strip())
    except Exception as e:
        print("highlight 텍스트 파일 읽어오기 실패 (이유 : {e}")

    out_cnt = 0
    redFont = Font(color="FF0000")

    for i, row in enumerate(ws.iter_rows(min_row = 2)):
        
        in_flag = False
        if row[0].value is None:
            break
        
        for word in keywords:
            if row[4].value.find(word) != -1:
                in_flag = True
                out_cnt += 1
                break

        if in_flag:
            row[4].font = redFont

    print("====================================================")
    print("                    필터링 결과")
    print("   총 %d 개의 키워드가 발견되어 하이라이팅 됐습니다" %(out_cnt))
    print("====================================================")


wb.save(excel_name)
wb.close()