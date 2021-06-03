from openpyxl import workbook, load_workbook, utils
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import requests
from bs4 import BeautifulSoup
import sys

print("\n=======================================")
print("            이메일 크롤링")
print("=======================================")

excel_name = "original.xlsx"

wb = load_workbook(excel_name, data_only=True)

excel_sheet_name = "Sheet1"  #작업할 시트 이름
start_row_value = "F2"       #작업할 셀 시작 번호
end_row_value = "F50"        #작업할 셀 마지막 번호
save_file_per = 5            #저장할 단위 지정

ws = wb[excel_sheet_name]
get_cells = ws[start_row_value : end_row_value]

isCoupang = 0 ; stopCoupang = 0 ; notCoupang = 0
total = 0

for row in get_cells:
    try:
        for cell in row:
            url = cell.value
            response = requests.get(url)
            
            if response.status_code == 200:
                html = response.text

                soup = BeautifulSoup(html, "html.parser")

                script = soup.script
                
                textScript = script.prettify()
                findTargetUrl = textScript.split("targetUrl")
                prettifyTargetUrl = findTargetUrl[1].split(";")
                prettifyAgain = prettifyTargetUrl[0].split('"')
                targetUrl = prettifyAgain[1]

                if(targetUrl.count("coupang") == 0 ) :
                    ws["G"+str(cell.row)] = "-"
                    notCoupang+=1
                    continue
                
                headers = {'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1 Safari/605.1.15'}

                targetResponse = requests.get(targetUrl,headers=headers)

                if targetResponse.status_code == 200:
                    html = targetResponse.text
                    soup = BeautifulSoup(html,"html.parser")
                    isStoppedSelling = soup.select('#contents > div.prod-atf.invalid-product > div > div.prod-buy > div.prod-not-find-known__buy__info > span.prod-not-find-known__buy__info__txt')
                    if len(isStoppedSelling) == 1 : 
                        ws["G"+str(cell.row)] = "배송/교환/반품/안내 란이 안떠서 이메일을 크롤링 할 수 없습니다."
                        stopCoupang+=1
                        continue

                    else :
                        isCoupang+=1

                    
                else:
                    print(targetResponse.status_code) 
            
            else:
                print(response.status_code) 
        
        if(total % save_file_per == 0) :
            wb.save(excel_name)
        total += 1
        print("[작업중]",row)


    except KeyboardInterrupt:
        print("KeyboardInterrupt")
        print("사용자의 요청으로 중단되었습니다. 현재까지 진행된 구역 : ",cell)
        wb.save(excel_name)
        wb.close()
        print("===================================================")
        print("쿠팡 링크 :",isCoupang,"/",total)
        print("중지된 쿠팡 링크 :",stopCoupang,"/",total)
        print("쿠팡 아닌 링크 :",notCoupang,"/",total)
        print("===================================================")
        sys.exit()

print("===================================================")
print("쿠팡 링크 :",isCoupang,"/",total)
print("중지된 쿠팡 링크 :",stopCoupang,"/",total)
print("쿠팡 아닌 링크 :",notCoupang,"/",total)
print("===================================================")
wb.save(excel_name)
wb.close()