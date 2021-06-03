'''시작전 기본 세팅
모듈 설치
pip install openxyl
pip install requests
pip install beautifulsoup4

크롤링 하는 동안 엑셀 파일은 열지 말아주세요. 열려있는 경우 결과값이 저장되지 않습니다.
'''
from openpyxl import Workbook, load_workbook, utils
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import requests
from bs4 import BeautifulSoup
import sys
import time

excel_file_name = "input"    #작업할 엑셀 파일 이름
excel_sheet_name = "Sheet1"  #작업할 엑셀 파일 시트 이름
start_row_value = "F2"      #작업할 셀 시작 번호
end_row_value = "F30"        #작업할 셀 마지막 번호
save_file_per = 5 #저장할 단위 지정

excel_file_name = excel_file_name+".xlsx"
wb = load_workbook(excel_file_name,data_only=True)
ws = wb[excel_sheet_name]
get_cells = ws[start_row_value : end_row_value]
totalFail = 0; total=0

start = time.time()

for row in get_cells:
    try:
        for cell in row:  
            total+=1
            url = cell.value
            response = requests.get(url)
            if(response.url.count("smartstore.naver.com")==0):
                print("[F A I L]",cell,"스마트 스토어 링크가 아닙니다.")
                totalFail+=1
                continue

            if response.status_code == 200:
                html = response.text
                soup = BeautifulSoup(html, 'html.parser')
                
                if(soup.title.text=="판매자의 사정에 따라 일시적으로 운영이 중지되었습니다."):
                    print("[F A I L]",cell,"판매자의 사정에 따라 일시적으로 운영이 중지되었습니다.")
                    totalFail+=1
                    continue
                
                if(soup.title.text=="상품이 존재하지 않습니다."):
                    print("[F A I L]",cell,"상품이 존재하지 않습니다.")
                    totalFail+=1
                    continue

                else:
                    try:
                        photoData = soup.select('#content > div > div._2-I30XS1lA > div._25tOXGEYJa > div._38rEjARje3 > div._23RpOU6xpc > img')
                        resPhotoSrc = photoData[0].attrs['data-src']

                        pTitle = soup.select('#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div._1ziwSSdAv8 > div.CxNYUPvHfB > h3')
                        resTitleText = ILLEGAL_CHARACTERS_RE.sub(r'',str(pTitle[0].text).strip())
                      
                        headData = soup.head.script
                        headDetail = headData.prettify()
                        hDatas = headDetail.split('productID')
                        hDDatas = hDatas[1].split('"')
                        productID = hDDatas[2]
                    
                        pID = '"id":"' + productID + '"'
                        rdFlag=0; oPFlag=0; sPFlag=0; csFlag=0; rvFlag=0

                        scriptOriginal = soup.select('body > script:nth-child(2)')
                        scriptText = scriptOriginal[0].prettify()
                        
                        scriptNarrowed=scriptText.split('exhibition')
                        for sN in scriptNarrowed:
                            if(sN.count(pID)):
                                sNParsed = sN.split('{')
                                for sNP in sNParsed:
                                    #제품 등록일
                                    if(sNP.count('"regDate"')):
                                        sNPParsed = sNP.split(',')
                                        for foundData in sNPParsed:
                                            if(foundData.count('"regDate"') and rdFlag==0):
                                                foundData = foundData.split(':')
                                                resRegisterDate = foundData[1][1:-3]
                                                rdFlag=1

                                    # 제품 원가 (할인 전 가격)
                                    # if(sNP.count('"salePrice"')):
                                    #     sNPParsed = sNP.split(',')
                                    #     for foundData in sNPParsed:
                                    #         if(foundData.count('"salePrice"') and oPFlag==0):
                                    #             foundData = foundData.split(':')
                                    #             # 엑셀에 저장하시려면 column 숫자 행번호에 맞게 수정하시면 됩니다. (행번호는 1부터 시작)
                                    #             resOriginalPrice = foundData[1]
                                    #             oPFlag=1
                                            
                                    #제품 세일 후 최종 가격
                                    if(sNP.count('"mobileDiscountedSalePrice"')):
                                        sNPParsed = sNP.split(',')
                                        for foundData in sNPParsed:
                                            if(foundData.count('"discountedSalePrice"') and sPFlag==0):
                                                foundData = foundData.split(':')
                                                resFinalSalePrice = foundData[1]
                                                sPFlag=1

                                    #구매 건수, 최근 구매 건수
                                    if(sNP.count('"cumulationSaleCount"') and csFlag==0):
                                        sNPParsed = sNP.split(',')
                                        cumulationSaleCount = sNPParsed[0].split(':')
                                        
                                        #구매 건수
                                        if(cumulationSaleCount[1] == '0'):
                                            resCumulationSaleCount = ""
                                        else :
                                            resCumulationSaleCount = cumulationSaleCount[1]
                                        recentSaleCount = sNPParsed[1].split(':')

                                        #최근 구매 건수
                                        if(recentSaleCount[1][:-1] == '0'):
                                            resRecentSaleCount =""
                                        else :
                                            resRecentSaleCount = recentSaleCount[1][:-1]
                                        csFlag=1 

                                    #전체 리뷰 수, 전체 리뷰 평점
                                    if(sNP.count('totalReviewCount') and rvFlag==0):
                                        sNPParsed = sNP.split(',')
                                        for foundData in sNPParsed:
                                            #전체 리뷰 수
                                            if(foundData.count('"totalReviewCount"')):
                                                finalData = foundData.split(':')
                                                if(finalData[1] == '0'):
                                                    resTotalReviewCount = ""
                                                else :
                                                    resTotalReviewCount = finalData[1]
                                    
                                            #전체 리뷰 평점
                                            if(foundData.count('"averageReviewScore"')):
                                                finalData = foundData.split(':')
                                                
                                                if(finalData[1] == '0'):
                                                    resTotalReviewScore = ""
                                                else :
                                                    resTotalReviewScore = finalData[1]
                                        rvFlag=1
                        
                        ws.cell(row = cell.row, column = 7, value=resPhotoSrc)
                        ws.cell(row = cell.row, column = 5, value=resTitleText)
                        ws.cell(row = cell.row, column = 3, value=resRegisterDate)
                        #제품 원가격(할인전) ws.cell(row = cell.row, column = 9999, value=resOriginalPrice)
                        ws.cell(row = cell.row, column = 13, value=resFinalSalePrice)
                        ws.cell(row = cell.row, column = 15, value=resCumulationSaleCount)
                        ws.cell(row = cell.row, column = 16, value=resRecentSaleCount)
                        ws.cell(row = cell.row, column = 17, value=resTotalReviewCount)
                        ws.cell(row = cell.row, column = 18, value=resTotalReviewScore)

                    except KeyboardInterrupt:
                        totalFail += 1
                        ws.cell(row = cell.row, column = 7, value=resPhotoSrc)
                        ws.cell(row = cell.row, column = 5, value=resTitleText)
                        ws.cell(row = cell.row, column = 3, value=resRegisterDate)
                        #제품 원가격(할인전) ws.cell(row = cell.row, column = 9999, value=resOriginalPrice)
                        ws.cell(row = cell.row, column = 13, value=resFinalSalePrice)
                        ws.cell(row = cell.row, column = 15, value=resCumulationSaleCount)
                        ws.cell(row = cell.row, column = 16, value=resRecentSaleCount)
                        ws.cell(row = cell.row, column = 17, value=resTotalReviewCount)
                        ws.cell(row = cell.row, column = 18, value=resTotalReviewScore)
                        print("KeyboardInterrupt")
                        print("[STOP]",cell)
                        wb.save(excel_file_name)
                        wb.close()
                        print("--------------------------------------------------")
                        print("Success / Total : %d / %d" %(total-totalFail,total))
                        print("Total Time : ",time.time() - start)    
                        sys.exit()                
                    
                    if(total%save_file_per==0):
                        wb.save(excel_file_name)

                    #셀 하나 처리 완료될때마다 status확인 하고 싶으면 주석 삭제하시면 됩니다.
                    # print("[SUCCESS]",cell)
                   
            else:
                print(response.status_code)     
    except KeyboardInterrupt:
        totalFail += 1
        print("KeyboardInterrupt")
        print("[STOP]",cell)
        wb.save(excel_file_name)
        wb.close()
        print("--------------------------------------------------")
        print("Success / Total : %d / %d" %(total-totalFail,total))
        print("Total Time : ",time.time() - start)    
        sys.exit()    

wb.save(excel_file_name)
wb.close()

print("--------------------------------------------------")
print("Success / Total : %d / %d" %(total-totalFail,total))
print("Total Time : ",time.time() - start)